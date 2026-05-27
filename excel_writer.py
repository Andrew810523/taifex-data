import os
from copy import copy

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Microsoft JhengHei")
DATA_FONT = Font(name="Microsoft JhengHei")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _display_width(s):
    if s is None:
        return 0
    w = 0
    for ch in str(s):
        w += 2 if ord(ch) > 127 else 1
    return w


# 這些欄位無論大小，一律以兩位小數顯示。
ALWAYS_TWO_DECIMALS = {
    "加權指數",
    "加權指數成交量(億)",
    "融資金額(億)",
    "融資金額變化量(億)",
}

# 這些欄位以三位小數顯示。
ALWAYS_THREE_DECIMALS = {
    "外資買方買權/賣權比",
    "外資買權/賣權比",
    "自營買方買權/賣權比",
    "散戶買/賣權比",
}

# 這些欄位顯示「0.XX%」(原樣加 % 字尾，不再乘 100)。
ALWAYS_RATIO_PERCENT = {
    "外資BC/SC增幅比例",
    "外資BP/SP增幅比例",
    "自營BC/SC增幅比例",
    "自營BP/SP增幅比例",
}


def _to_cell_value(v):
    """Convert NaN/NaT/None to Python None so openpyxl writes a blank cell."""
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    return v


def write_df_preserve_style(file_path: str, sheet_name: str, df: pd.DataFrame) -> bool:
    """Write df into sheet_name preserving all existing cell styling.

    - If the sheet exists: overlays values onto existing cells; styles (font, fill,
      alignment, column widths, freeze panes, number_format) are kept as-is.
    - If the sheet doesn't exist: creates it and applies the default style once.

    Returns True if the sheet was newly created, False if it already existed.
    """
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
        # remove default empty sheet
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            del wb["Sheet"]

    newly_created = sheet_name not in wb.sheetnames
    if newly_created:
        ws = wb.create_sheet(title=sheet_name)
    else:
        ws = wb[sheet_name]

    n_rows = len(df)
    n_cols = len(df.columns)
    first_col_name = df.columns[0]

    # 偵測標題列佈局：使用者可能在頂端加了「群組標題」一列。
    # - 單列標題: A1 = "日期"，資料從 row 2 開始
    # - 雙列標題: A1 = 群組(None/"大盤資訊"等)、A2 = "日期"，資料從 row 3 開始
    if newly_created:
        header_row = 1
        data_start = 2
    else:
        a1 = ws.cell(row=1, column=1).value
        a2 = ws.cell(row=2, column=1).value
        if a2 == first_col_name and a1 != first_col_name:
            header_row = 2  # 雙列標題：row 1 群組保留不動
            data_start = 3
        else:
            header_row = 1
            data_start = 2

    # 寫欄名 (僅在 header_row=1 時；雙列標題的 row 2 保留使用者欄名)
    if header_row == 1:
        for c_idx, col in enumerate(df.columns, start=1):
            ws.cell(row=1, column=c_idx, value=col)

    # 寫資料
    for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=data_start):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=_to_cell_value(val))

    # 不刪除任何舊列、不刪除右側欄位 — 保留使用者既有資料/格式

    if newly_created:
        apply_style_to_worksheet(ws)

    wb.save(file_path)
    return newly_created


def _date_to_str(v):
    """Excel cell 日期值轉成 'YYYY-MM-DD' 字串 (datetime / date / str 都接受)。"""
    if v is None:
        return None
    import datetime as _dt
    if isinstance(v, _dt.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, _dt.date):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return s[:10] if s else None


def update_rows_by_date(file_path: str, sheet_name: str, rows_dict: dict):
    """只新增/替換指定日期的列，其他列完全不動。
    rows_dict: {date_str: {col_name: value, ...}}
    - 既有相同日期 → 覆寫該列指定欄位的「值」(其他欄不動、樣式保留)
    - 新日期 → append 到 max_row 之下
    - 不刪除任何舊資料、不改 header
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(file_path)
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"sheet '{sheet_name}' 不存在")
    ws = wb[sheet_name]

    # 偵測 header row (1 或 2)
    a1 = ws.cell(row=1, column=1).value
    a2 = ws.cell(row=2, column=1).value
    if a1 == "日期":
        header_row_idx = 1
    elif a2 == "日期":
        header_row_idx = 2
    else:
        raise ValueError(f"找不到「日期」欄 (A1={a1!r}, A2={a2!r})")

    # 建立欄名 -> 欄索引
    col_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row_idx, column=c).value
        if v is not None:
            col_map[v] = c

    # 建立既有日期 -> row 索引
    date_col_idx = col_map.get("日期", 1)
    date_to_row = {}
    for r in range(header_row_idx + 1, ws.max_row + 1):
        d_str = _date_to_str(ws.cell(row=r, column=date_col_idx).value)
        if d_str:
            date_to_row[d_str] = r

    next_append_row = ws.max_row + 1
    # 上一列 (作為新增列的樣式來源)
    template_row = ws.max_row if ws.max_row > header_row_idx else None

    def _copy_style(src_cell, dst_cell):
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.border = copy(src_cell.border)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy(src_cell.protection)

    for date_str, row_values in rows_dict.items():
        if date_str in date_to_row:
            target_row = date_to_row[date_str]
            is_new_row = False
        else:
            target_row = next_append_row
            next_append_row += 1
            is_new_row = True
            # 寫入日期 (新增列)
            ws.cell(row=target_row, column=date_col_idx, value=date_str)
            # 從前一列複製所有欄位的樣式 (字型/對齊/邊框/number_format 等)
            if template_row is not None:
                for c in range(1, ws.max_column + 1):
                    _copy_style(ws.cell(row=template_row, column=c), ws.cell(row=target_row, column=c))

        for col_name, val in row_values.items():
            col_idx = col_map.get(col_name)
            if col_idx is None:
                continue  # 該分頁沒有這欄，跳過
            if col_name == "日期":
                continue  # 已經寫過
            ws.cell(row=target_row, column=col_idx, value=_to_cell_value(val))

    wb.save(file_path)


def detect_header_row(file_path: str, sheet_name: str, expected_first_col: str = "日期") -> int:
    """回傳 pandas read_excel 用的 header 參數 (0 = row1 是欄名；1 = row2 是欄名)。
    若分頁不存在或找不到 expected_first_col，回傳 0。"""
    if not os.path.exists(file_path):
        return 0
    wb = load_workbook(file_path, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return 0
    ws = wb[sheet_name]
    a1 = ws.cell(row=1, column=1).value
    a2 = ws.cell(row=2, column=1).value
    wb.close()
    if a1 == expected_first_col:
        return 0
    if a2 == expected_first_col:
        return 1
    return 0


def append_row(file_path: str, row_dict: dict, columns_order: list):
    """Insert / update a single row (by 日期) in the first sheet.
    既有資料完全不動，只動該日期那一列；新日期則 append。"""
    file_exists = os.path.exists(file_path)
    date_val = row_dict.get("日期")
    if date_val is None:
        raise ValueError("row_dict 必須含「日期」")

    # 確保欄位順序符合 columns_order (補空)
    full_row = {col: row_dict.get(col) for col in columns_order}

    if not file_exists:
        # 建立新檔 + 預設樣式
        first_sheet = "爬蟲資料"
        df = pd.DataFrame([full_row], columns=columns_order)
        write_df_preserve_style(file_path, first_sheet, df)
        return

    wb_existing = load_workbook(file_path, read_only=True)
    first_sheet = wb_existing.sheetnames[0]
    wb_existing.close()

    update_rows_by_date(file_path, first_sheet, {date_val: full_row})


def apply_style_to_worksheet(ws):
    """Apply default style to a worksheet (used only on initial sheet creation)."""
    headers = {cell.column_letter: cell.value for cell in ws[1]}
    two_dec_cols = {col for col, h in headers.items() if h in ALWAYS_TWO_DECIMALS}
    three_dec_cols = {col for col, h in headers.items() if h in ALWAYS_THREE_DECIMALS}
    ratio_pct_cols = {col for col, h in headers.items() if h in ALWAYS_RATIO_PERCENT}

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = CENTER
            cell.font = DATA_FONT
            cell.border = BORDER
            if cell.column_letter in two_dec_cols and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
            elif cell.column_letter in three_dec_cols and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.000"
            elif cell.column_letter in ratio_pct_cols and isinstance(cell.value, (int, float)):
                cell.number_format = '0.00"%"'
            elif isinstance(cell.value, float):
                if cell.column_letter == "A":
                    pass
                elif abs(cell.value) >= 1000:
                    cell.number_format = "#,##0"
                else:
                    cell.number_format = "#,##0.00"
            elif isinstance(cell.value, int):
                cell.number_format = "#,##0"

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_w = 0
        for cell in col_cells:
            v = cell.value
            if isinstance(v, float):
                v = f"{v:,.2f}" if abs(v) < 1000 else f"{v:,.0f}"
            elif isinstance(v, int):
                v = f"{v:,}"
            max_w = max(max_w, _display_width(v))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(max_w + 2, 10)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
