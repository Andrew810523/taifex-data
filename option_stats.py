"""選擇權 分頁計算與寫入。

公式：
- 增幅 = 今日 − 昨日
- 外資(SC增幅-BC增幅)金額 = (T_SC − Y_SC) − (T_BC − Y_BC)
- 外資BC/SC增幅比例 = BC增幅 / SC增幅
- (BP/SP 同理；自營同理)
"""
import os

import pandas as pd

import excel_writer
import stats as stats_module

OPT_SHEET = "選擇權"

OPT_COLS = [
    "日期",
    "結算日",
    "當次月大台留倉差增減",
    "加權指數",
    "加權指數漲跌",
    "加權指數漲跌幅(%)",
    "外資BC口數", "外資BC金額", "外資SC口數", "外資SC金額",
    "外資(SC增幅-BC增幅)金額", "外資BC/SC增幅比例",
    "外資BP口數", "外資BP金額", "外資SP口數", "外資SP金額",
    "外資(SP增幅-BP增幅)金額", "外資BP/SP增幅比例",
    "自營BC口數", "自營BC金額", "自營SC口數", "自營SC金額",
    "自營(SC增幅-BC增幅)金額", "自營BC/SC增幅比例",
    "自營BP口數", "自營BP金額", "自營SP口數", "自營SP金額",
    "自營(SP增幅-BP增幅)金額", "自營BP/SP增幅比例",
]


def _safe_sub(a, b):
    if a is None or b is None or pd.isna(a) or pd.isna(b):
        return None
    return a - b


def _safe_div(a, b):
    if a is None or b is None or pd.isna(a) or pd.isna(b) or b == 0:
        return None
    return a / b


def _compute(raw: pd.DataFrame) -> pd.DataFrame:
    df = raw.copy()
    df["_d"] = df["日期"].apply(stats_module._to_date)
    df = df.sort_values("_d").reset_index(drop=True)
    df["_當次月大台差"] = df["當月大台未沖銷契約量"] - df["次月大台未沖銷契約量"]

    out_rows = []
    for i, row in df.iterrows():
        prev = df.iloc[i - 1] if i > 0 else None
        d = row["_d"]

        def inc(col):
            if prev is None:
                return None
            return _safe_sub(row.get(col), prev.get(col))

        def growth_ratio(col):
            """ΔX / X_y (今日相對昨日的成長率)。"""
            if prev is None:
                return None
            y = prev.get(col)
            t = row.get(col)
            if y is None or t is None or pd.isna(y) or pd.isna(t) or y == 0:
                return None
            return (t - y) / y

        e_BC, e_SC = inc("外資BC金額"), inc("外資SC金額")
        e_BP, e_SP = inc("外資BP金額"), inc("外資SP金額")
        d_BC, d_SC = inc("自營BC金額"), inc("自營SC金額")
        d_BP, d_SP = inc("自營BP金額"), inc("自營SP金額")
        # 成長率 (用於 增幅比例)
        e_BC_g, e_SC_g = growth_ratio("外資BC金額"), growth_ratio("外資SC金額")
        e_BP_g, e_SP_g = growth_ratio("外資BP金額"), growth_ratio("外資SP金額")
        d_BC_g, d_SC_g = growth_ratio("自營BC金額"), growth_ratio("自營SC金額")
        d_BP_g, d_SP_g = growth_ratio("自營BP金額"), growth_ratio("自營SP金額")

        out = {
            "日期": str(d),
            "結算日": stats_module._settlement_label(d),
            "當次月大台留倉差增減": (
                _safe_sub(row["_當次月大台差"], prev["_當次月大台差"]) if prev is not None else None
            ),
            "加權指數": round(row["加權指數"], 2) if pd.notna(row.get("加權指數")) else None,
            "加權指數漲跌": row.get("加權指數漲跌"),
            "加權指數漲跌幅(%)": row.get("加權指數漲跌幅"),
            "外資BC口數": row.get("外資BC口數"),
            "外資BC金額": row.get("外資BC金額"),
            "外資SC口數": row.get("外資SC口數"),
            "外資SC金額": row.get("外資SC金額"),
            "外資(SC增幅-BC增幅)金額": _safe_sub(e_SC, e_BC),
            "外資BC/SC增幅比例": _safe_div(e_BC_g, e_SC_g),
            "外資BP口數": row.get("外資BP口數"),
            "外資BP金額": row.get("外資BP金額"),
            "外資SP口數": row.get("外資SP口數"),
            "外資SP金額": row.get("外資SP金額"),
            "外資(SP增幅-BP增幅)金額": _safe_sub(e_SP, e_BP),
            "外資BP/SP增幅比例": _safe_div(e_BP_g, e_SP_g),
            "自營BC口數": row.get("自營BC口數"),
            "自營BC金額": row.get("自營BC金額"),
            "自營SC口數": row.get("自營SC口數"),
            "自營SC金額": row.get("自營SC金額"),
            "自營(SC增幅-BC增幅)金額": _safe_sub(d_SC, d_BC),
            "自營BC/SC增幅比例": _safe_div(d_BC_g, d_SC_g),
            "自營BP口數": row.get("自營BP口數"),
            "自營BP金額": row.get("自營BP金額"),
            "自營SP口數": row.get("自營SP口數"),
            "自營SP金額": row.get("自營SP金額"),
            "自營(SP增幅-BP增幅)金額": _safe_sub(d_SP, d_BP),
            "自營BP/SP增幅比例": _safe_div(d_BP_g, d_SP_g),
        }
        out_rows.append(out)

    return pd.DataFrame(out_rows, columns=OPT_COLS)


def recalculate(file_path: str, target_dates=None):
    """從第一個 sheet 重算選擇權。預設只 upsert target_dates 指定的日期列，
    其他歷史列完全不動 (保留使用者手動編輯)。target_dates=None 時改寫整張表。"""
    if not os.path.exists(file_path):
        raise FileNotFoundError(file_path)

    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True)
    first_sheet = wb.sheetnames[0]
    wb.close()
    raw_header = excel_writer.detect_header_row(file_path, first_sheet, expected_first_col="日期")
    raw = pd.read_excel(file_path, sheet_name=first_sheet, dtype={"日期": str}, header=raw_header)
    if raw.empty:
        return

    df = _compute(raw)

    if target_dates is not None:
        target_set = set(target_dates)
        df_target = df[df["日期"].isin(target_set)]
        rows_dict = {row["日期"]: row.to_dict() for _, row in df_target.iterrows()}
        if rows_dict:
            excel_writer.update_rows_by_date(file_path, OPT_SHEET, rows_dict)
    else:
        excel_writer.write_df_preserve_style(file_path, OPT_SHEET, df)


if __name__ == "__main__":
    import sys
    target = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "籌碼資料.xlsx"
    )
    recalculate(target)
    print(f"選擇權分頁已更新: {target}")
