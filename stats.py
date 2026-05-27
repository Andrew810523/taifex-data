import os
from datetime import date, datetime, timedelta

import pandas as pd
from openpyxl import load_workbook

import excel_writer

STATS_SHEET = "統計"

STATS_COLS = [
    "日期",
    "結算日",
    "當次月大台留倉差增減",
    "加權指數",
    "加權指數漲跌",
    "加權指數當日高低差",
    "加權指數漲跌幅(%)",
    "加權指數成交量(億)",
    "融資金額(億)",
    "融資金額變化量(億)",
    "融券(交易單位)",
    "融券單位變化量",
    "PCR與結算比",
    "外資現貨買賣超",
    "外資(大小台)期貨未平倉",
    "外資期貨未平倉與結算比",
    "外資期貨未平倉與前日增減",
    "外資(買)OP未平倉口數",
    "外資(買)OP未平倉金額",
    "外資(BC)OP未平倉金額與結算比",
    "外資(賣)OP未平倉口數",
    "外資(賣)OP未平倉金額",
    "外資(BP)OP未平倉金額與結算比",
    "外資買方買權/賣權比",
    "外資買權/賣權比",
    "自營(自行)現貨買賣超",
    "自營(避險)現貨買賣超",
    "自營(大小台)期貨未平倉",
    "自營(買)OP未平倉口數",
    "自營(買)OP未平倉金額",
    "自營(BC)OP未平倉金額與結算比",
    "自營(賣)OP未平倉口數",
    "自營(賣)OP未平倉金額",
    "自營(BP)OP未平倉金額與結算比",
    "自營買方買權/賣權比",
    "散戶買/賣權比",
    "散戶看多",
    "散戶看空",
    "當次月小台留倉差增減",
    "散戶未平倉",
    "微台散戶看多",
    "微台散戶看空",
    "當次月微台留倉差增減",
    "微台散戶未平倉",
    "微台散戶多空比",
]

# 已知/已公告的台灣加權股市休市日（含原應為結算日但遇假期順延的情況）。
# 結算日預設為每月第三個星期三，遇下列日期則順延至下一個交易日。
# 需要時可手動補上新年度的假期。
_HOLIDAYS = {
    # 2022
    date(2022, 1, 31), date(2022, 2, 1), date(2022, 2, 2), date(2022, 2, 3), date(2022, 2, 4),
    date(2022, 2, 28),
    date(2022, 4, 4), date(2022, 4, 5),
    date(2022, 5, 2),
    date(2022, 6, 3),
    date(2022, 9, 9),
    date(2022, 10, 10),
    # 2023
    date(2023, 1, 2),
    date(2023, 1, 20), date(2023, 1, 23), date(2023, 1, 24), date(2023, 1, 25), date(2023, 1, 26), date(2023, 1, 27),
    date(2023, 2, 27), date(2023, 2, 28),
    date(2023, 4, 3), date(2023, 4, 4), date(2023, 4, 5),
    date(2023, 5, 1),
    date(2023, 6, 22), date(2023, 6, 23),
    date(2023, 9, 29),
    date(2023, 10, 9), date(2023, 10, 10),
    # 2024
    date(2024, 1, 1),
    date(2024, 2, 8), date(2024, 2, 9), date(2024, 2, 12), date(2024, 2, 13), date(2024, 2, 14),
    date(2024, 2, 28),
    date(2024, 4, 4), date(2024, 4, 5),
    date(2024, 5, 1),
    date(2024, 6, 10),
    date(2024, 9, 17),
    date(2024, 10, 10),
    # 2025
    date(2025, 1, 1),
    date(2025, 1, 27), date(2025, 1, 28), date(2025, 1, 29), date(2025, 1, 30), date(2025, 1, 31),
    date(2025, 2, 28),
    date(2025, 4, 3), date(2025, 4, 4),
    date(2025, 5, 1),
    date(2025, 5, 30),
    date(2025, 10, 6),
    date(2025, 10, 10),
    # 2026
    date(2026, 1, 1),
    date(2026, 2, 16), date(2026, 2, 17), date(2026, 2, 18), date(2026, 2, 19), date(2026, 2, 20),
    date(2026, 2, 27),
    date(2026, 4, 3), date(2026, 4, 6),
    date(2026, 5, 1),
    date(2026, 6, 19),
    date(2026, 9, 25),
    date(2026, 10, 9),
}


def _to_date(d):
    if isinstance(d, datetime):
        return d.date()
    if isinstance(d, date):
        return d
    if isinstance(d, str):
        return datetime.strptime(d[:10], "%Y-%m-%d").date()
    raise ValueError(f"Unsupported date value: {d!r}")


def _next_business_day(d: date) -> date:
    nxt = d + timedelta(days=1)
    while nxt.weekday() >= 5 or nxt in _HOLIDAYS:
        nxt += timedelta(days=1)
    return nxt


def _settlement_day_of_month(year: int, month: int) -> date:
    # 每月第三個星期三
    first = date(year, month, 1)
    offset = (2 - first.weekday()) % 7  # weekday: Mon=0, Wed=2
    third_wed = first + timedelta(days=offset + 14)
    if third_wed in _HOLIDAYS or third_wed.weekday() >= 5:
        third_wed = _next_business_day(third_wed)
    return third_wed


def _is_settlement_day(d: date) -> bool:
    return d == _settlement_day_of_month(d.year, d.month)


def _settlement_label(d: date) -> int:
    if _is_settlement_day(d):
        return 999
    return d.weekday() + 1  # Mon=1 ... Sun=7


def _safe_sub(a, b):
    if pd.isna(a) or pd.isna(b):
        return None
    return a - b


def _safe_div(a, b):
    if pd.isna(a) or pd.isna(b) or b == 0:
        return None
    return a / b


def _equiv_from_raw(row):
    """從原始 sheet 列抽出與「統計欄位」對應的關鍵指標，供跨日 / 與結算日比較使用。"""
    return {
        "融資金額(億)": row["融資金額"] / 1e5 if pd.notna(row["融資金額"]) else None,
        "融券(交易單位)": row["融券"],
        "外資(大小台)期貨未平倉": (
            row["外資大台期貨未平倉"]
            + row["外資小台期貨未平倉"] / 4
            + (row.get("外資微台期貨未平倉") or 0) / 20
        ),
        "外資(買)OP未平倉金額": row["外資BC金額"] - row["外資SC金額"],
        "外資(賣)OP未平倉金額": row["外資BP金額"] - row["外資SP金額"],
        "自營(買)OP未平倉金額": row["自營BC金額"] - row["自營SC金額"],
        "自營(賣)OP未平倉金額": row["自營BP金額"] - row["自營SP金額"],
        # BC/BP 金額直接值 (給「與結算比」用，不是淨額)
        "外資BC金額": row.get("外資BC金額"),
        "外資BP金額": row.get("外資BP金額"),
        "自營BC金額": row.get("自營BC金額"),
        "自營BP金額": row.get("自營BP金額"),
        "PCR未平倉比": row.get("PCR未平倉比"),
    }


def _equiv_from_stats(row):
    """從統計分頁抽出對應指標 (給 raw 缺前一日 / 結算日時當 fallback)。注意統計沒存原始 PCR。"""
    return {
        "融資金額(億)": row.get("融資金額(億)"),
        "融券(交易單位)": row.get("融券(交易單位)"),
        "外資(大小台)期貨未平倉": row.get("外資(大小台)期貨未平倉"),
        "外資(買)OP未平倉金額": row.get("外資(買)OP未平倉金額"),
        "外資(賣)OP未平倉金額": row.get("外資(賣)OP未平倉金額"),
        "自營(買)OP未平倉金額": row.get("自營(買)OP未平倉金額"),
        "自營(賣)OP未平倉金額": row.get("自營(賣)OP未平倉金額"),
        # 統計分頁未存原始 BC/BP 金額，fallback 拿不到
        "外資BC金額": None,
        "外資BP金額": None,
        "自營BC金額": None,
        "自營BP金額": None,
        "PCR未平倉比": None,
    }


def _build_stats_lookup(prior_stats: pd.DataFrame):
    """回傳 (sorted_dates, by_date_equiv, sorted_settle_dates, settle_equiv)。"""
    if prior_stats is None or prior_stats.empty:
        return [], {}, [], {}
    by_date = {}
    settle = {}
    for _, r in prior_stats.iterrows():
        if pd.isna(r.get("日期")):
            continue
        try:
            d = _to_date(r["日期"])
        except Exception:
            continue
        eq = _equiv_from_stats(r)
        by_date[d] = eq
        settle_flag = r.get("結算日")
        if pd.notna(settle_flag) and int(settle_flag) == 999:
            settle[d] = eq
    return sorted(by_date), by_date, sorted(settle), settle


def _compute(df_raw: pd.DataFrame, prior_stats: pd.DataFrame = None) -> pd.DataFrame:
    df = df_raw.copy()
    df["_d"] = df["日期"].apply(_to_date)
    df = df.sort_values("_d").reset_index(drop=True)

    # 讓所有可能缺欄位的欄位有預設值，避免 KeyError。
    for col in [
        "外資現貨買賣超", "PCR未平倉比", "外資微台期貨未平倉", "自營微台期貨未平倉",
    ]:
        if col not in df.columns:
            df[col] = float("nan")

    df["_外資大小台"] = (
        df["外資大台期貨未平倉"]
        + df["外資小台期貨未平倉"] / 4
        + df["外資微台期貨未平倉"].fillna(0) / 20
    )
    df["_外資BC_OP金額"] = df["外資BC金額"] - df["外資SC金額"]
    df["_外資BP_OP金額"] = df["外資BP金額"] - df["外資SP金額"]
    df["_自營大小台"] = (
        df["自營大台期貨未平倉"]
        + df["自營小台期貨未平倉"] / 4
        + df["自營微台期貨未平倉"].fillna(0) / 20
    )
    df["_自營BC_OP金額"] = df["自營BC金額"] - df["自營SC金額"]
    df["_自營BP_OP金額"] = df["自營BP金額"] - df["自營SP金額"]
    df["_當次月大台差"] = df["當月大台未沖銷契約量"] - df["次月大台未沖銷契約量"]
    df["_當次月小台差"] = df["當月小台未沖銷契約量"] - df["次月小台未沖銷契約量"]
    df["_當次月微台差"] = df["當月微台未沖銷契約量"] - df["次月微台未沖銷契約量"]
    df["_是結算日"] = df["_d"].apply(_is_settlement_day)

    raw_dates = set(df["_d"])
    sorted_stat_dates, stats_by_date, sorted_stat_settle, stats_settle = _build_stats_lookup(prior_stats)

    def _yesterday_equiv(today: date, prev_raw_row):
        if prev_raw_row is not None:
            return _equiv_from_raw(prev_raw_row)
        # raw 沒前一日 → 從統計中抽「< today 的最後一筆」(且不是 raw 自己)
        for d in reversed(sorted_stat_dates):
            if d < today and d not in raw_dates:
                return stats_by_date[d]
        return None

    def _anchor_equiv(today: date, anchor_raw_row):
        if anchor_raw_row is not None:
            return _equiv_from_raw(anchor_raw_row)
        for d in reversed(sorted_stat_settle):
            if d <= today and d not in raw_dates:
                return stats_settle[d]
        return None

    out_rows = []
    last_settle_idx = None
    for i, row in df.iterrows():
        d = row["_d"]
        prev = df.iloc[i - 1] if i > 0 else None

        if row["_是結算日"]:
            anchor = row
        elif last_settle_idx is not None:
            anchor = df.iloc[last_settle_idx]
        else:
            anchor = None

        y_eq = _yesterday_equiv(d, prev)
        s_eq = _anchor_equiv(d, anchor)

        def y(col):
            return None if prev is None else prev[col]

        def s(col):
            return None if anchor is None else anchor[col]

        def y_eq_v(col):
            return None if y_eq is None else y_eq.get(col)

        def s_eq_v(col):
            return None if s_eq is None else s_eq.get(col)

        # 今日的指標 (raw)
        today_融資億 = row["融資金額"] / 1e5 if pd.notna(row["融資金額"]) else None
        today_外資大小台 = row["_外資大小台"]
        today_外資BC金 = row["_外資BC_OP金額"]
        today_外資BP金 = row["_外資BP_OP金額"]
        today_自營BC金 = row["_自營BC_OP金額"]
        today_自營BP金 = row["_自營BP_OP金額"]

        out = {
            "日期": str(d),
            "結算日": _settlement_label(d),
            "當次月大台留倉差增減": _safe_sub(row["_當次月大台差"], y("_當次月大台差")),
            "加權指數": round(row["加權指數"], 2) if pd.notna(row["加權指數"]) else None,
            "加權指數漲跌": row["加權指數漲跌"],
            "加權指數當日高低差": row["加權指數當日高低差"],
            "加權指數漲跌幅(%)": row["加權指數漲跌幅"],
            "加權指數成交量(億)": round(row["加權指數成交量"], 2) if pd.notna(row["加權指數成交量"]) else None,
            "融資金額(億)": round(today_融資億, 2) if today_融資億 is not None else None,
            "融資金額變化量(億)": (
                round(today_融資億 - y_eq_v("融資金額(億)"), 2)
                if today_融資億 is not None and y_eq_v("融資金額(億)") is not None
                else None
            ),
            "融券(交易單位)": row["融券"],
            "融券單位變化量": _safe_sub(row["融券"], y_eq_v("融券(交易單位)")),
            "PCR與結算比": _safe_sub(row["PCR未平倉比"], s_eq_v("PCR未平倉比")),
            "外資現貨買賣超": row["外資現貨買賣超"] / 1e8 if pd.notna(row["外資現貨買賣超"]) else None,
            "外資(大小台)期貨未平倉": today_外資大小台,
            "外資期貨未平倉與結算比": _safe_sub(today_外資大小台, s_eq_v("外資(大小台)期貨未平倉")),
            "外資期貨未平倉與前日增減": _safe_sub(today_外資大小台, y_eq_v("外資(大小台)期貨未平倉")),
            "外資(買)OP未平倉口數": _safe_sub(row["外資BC口數"], row["外資SC口數"]),
            "外資(買)OP未平倉金額": today_外資BC金,
            "外資(BC)OP未平倉金額與結算比": _safe_sub(row.get("外資BC金額"), s_eq_v("外資BC金額")),
            "外資(賣)OP未平倉口數": _safe_sub(row["外資BP口數"], row["外資SP口數"]),
            "外資(賣)OP未平倉金額": today_外資BP金,
            "外資(BP)OP未平倉金額與結算比": _safe_sub(row.get("外資BP金額"), s_eq_v("外資BP金額")),
            "外資買方買權/賣權比": (
                round(_safe_div(row["外資買方買權金額"], row["外資買方賣權金額"]), 3)
                if _safe_div(row["外資買方買權金額"], row["外資買方賣權金額"]) is not None else None
            ),
            "外資買權/賣權比": (
                round(_safe_div(today_外資BC金, today_外資BP金), 3)
                if _safe_div(today_外資BC金, today_外資BP金) is not None else None
            ),
            "自營(自行)現貨買賣超": row["自營(自行)現貨買賣超"] / 1e8 if pd.notna(row["自營(自行)現貨買賣超"]) else None,
            "自營(避險)現貨買賣超": row["自營(避險)現貨買賣超"] / 1e8 if pd.notna(row["自營(避險)現貨買賣超"]) else None,
            "自營(大小台)期貨未平倉": row["_自營大小台"],
            "自營(買)OP未平倉口數": _safe_sub(row["自營BC口數"], row["自營SC口數"]),
            "自營(買)OP未平倉金額": today_自營BC金,
            "自營(BC)OP未平倉金額與結算比": _safe_sub(row.get("自營BC金額"), s_eq_v("自營BC金額")),
            "自營(賣)OP未平倉口數": _safe_sub(row["自營BP口數"], row["自營SP口數"]),
            "自營(賣)OP未平倉金額": today_自營BP金,
            "自營(BP)OP未平倉金額與結算比": _safe_sub(row.get("自營BP金額"), s_eq_v("自營BP金額")),
            "自營買方買權/賣權比": (
                round(_safe_div(row["自營買方買權金額"], row["自營買方賣權金額"]), 3)
                if _safe_div(row["自營買方買權金額"], row["自營買方賣權金額"]) is not None else None
            ),
            # 散戶買/賣權比 = |(外資+自營 BC − SC) / (外資+自營 BP − SP)|，排投信、取絕對值、3 位小數
            "散戶買/賣權比": (
                lambda n, d: round(abs(n / d), 3) if (n is not None and d not in (None, 0)) else None
            )(
                _safe_sub(
                    (row.get("外資BC金額") or 0) + (row.get("自營BC金額") or 0),
                    (row.get("外資SC金額") or 0) + (row.get("自營SC金額") or 0),
                ) if pd.notna(row.get("外資BC金額")) else None,
                _safe_sub(
                    (row.get("外資BP金額") or 0) + (row.get("自營BP金額") or 0),
                    (row.get("外資SP金額") or 0) + (row.get("自營SP金額") or 0),
                ) if pd.notna(row.get("外資BP金額")) else None,
            ),
        }

        # 散戶 (小台)
        retail_long = _safe_sub(row["小台未沖銷契約量"], row["三大法人多方小台口數"])
        retail_short = _safe_sub(row["小台未沖銷契約量"], row["三大法人空方小台口數"])
        out["散戶看多"] = retail_long
        out["散戶看空"] = retail_short
        out["當次月小台留倉差增減"] = _safe_sub(row["_當次月小台差"], y("_當次月小台差"))
        out["散戶未平倉"] = (
            None if retail_long is None or retail_short is None else retail_long - retail_short
        )

        # 微台散戶
        m_long = _safe_sub(row["微台未沖銷契約量"], row["三大法人多方微台口數"])
        m_short = _safe_sub(row["微台未沖銷契約量"], row["三大法人空方微台口數"])
        out["微台散戶看多"] = m_long
        out["微台散戶看空"] = m_short
        out["當次月微台留倉差增減"] = _safe_sub(row["_當次月微台差"], y("_當次月微台差"))
        m_open = (
            None if m_long is None or m_short is None else m_long - m_short
        )
        out["微台散戶未平倉"] = m_open
        # 微台散戶多空比 = (微台散戶看多 − 微台散戶看空) / 微台未沖銷契約量 × 100 (%)
        tmf_total = row.get("微台未沖銷契約量")
        if m_long is None or m_short is None or pd.isna(tmf_total) or tmf_total == 0:
            out["微台散戶多空比"] = None
        else:
            out["微台散戶多空比"] = (m_long - m_short) / tmf_total * 100

        out_rows.append(out)

        if row["_是結算日"]:
            last_settle_idx = i

    return pd.DataFrame(out_rows, columns=STATS_COLS)


def _read_existing_stats(file_path: str) -> pd.DataFrame:
    """讀取現有的統計分頁；若不存在或為空則回傳空 DataFrame。
    自動偵測單列 vs 雙列標題（使用者可能加了群組標題列）。"""
    wb = load_workbook(file_path, read_only=True)
    has_sheet = STATS_SHEET in wb.sheetnames
    wb.close()
    if not has_sheet:
        return pd.DataFrame(columns=STATS_COLS)
    header_row = excel_writer.detect_header_row(file_path, STATS_SHEET, expected_first_col="日期")
    df = pd.read_excel(file_path, sheet_name=STATS_SHEET, dtype={"日期": str}, header=header_row)
    if df.empty or "日期" not in df.columns:
        return pd.DataFrame(columns=STATS_COLS)
    df["日期"] = df["日期"].astype(str).str.slice(0, 10)
    return df


def _merge(existing: pd.DataFrame, fresh: pd.DataFrame) -> pd.DataFrame:
    """以 fresh (從原始 sheet 重算) 覆蓋 existing 同日期的列。fresh 為 NaN 而 existing
    有值時，保留 existing 值 (用於：無法從 raw 算的欄位 + 使用者手動填補的 cell)。"""
    if existing.empty:
        merged = fresh
    else:
        for col in STATS_COLS:
            if col not in existing.columns:
                existing[col] = None
            if col not in fresh.columns:
                fresh[col] = None
        existing = existing[STATS_COLS]
        fresh = fresh[STATS_COLS].copy()

        existing_by_date = {row["日期"]: row for _, row in existing.iterrows()}
        for i, row in fresh.iterrows():
            ex = existing_by_date.get(row["日期"])
            if ex is None:
                continue
            for col in STATS_COLS:
                if pd.isna(row[col]) and pd.notna(ex[col]):
                    fresh.at[i, col] = ex[col]

        keep = existing[~existing["日期"].isin(set(fresh["日期"]))]
        merged = pd.concat([keep, fresh], ignore_index=True)

    merged["_d"] = merged["日期"].apply(_to_date)
    merged = merged.sort_values("_d").drop(columns=["_d"]).reset_index(drop=True)
    return merged[STATS_COLS]


def write_stats_sheet(file_path: str, df: pd.DataFrame):
    """寫入「統計」分頁，保留既有 cell 樣式（欄寬/凍結/標題色/格式都不動）。
    若分頁不存在則建立並套用一次預設樣式。"""
    excel_writer.write_df_preserve_style(file_path, STATS_SHEET, df)


def recalculate(file_path: str, target_dates=None):
    """從第一個 sheet 重算統計。預設只 upsert target_dates 指定的日期列，
    其他歷史列完全不動 (保留使用者手動編輯)。target_dates=None 時保持既有「合併」行為。
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(file_path)

    wb = load_workbook(file_path, read_only=True)
    first_sheet = wb.sheetnames[0]
    wb.close()
    raw_header = excel_writer.detect_header_row(file_path, first_sheet, expected_first_col="日期")
    raw = pd.read_excel(file_path, sheet_name=first_sheet, dtype={"日期": str}, header=raw_header)
    if raw.empty:
        return

    existing = _read_existing_stats(file_path)
    fresh = _compute(raw, prior_stats=existing)

    if target_dates is not None:
        # 只 upsert 指定日期，其他列完全不動
        target_set = set(target_dates)
        fresh_target = fresh[fresh["日期"].isin(target_set)]
        rows_dict = {row["日期"]: row.to_dict() for _, row in fresh_target.iterrows()}
        if rows_dict:
            excel_writer.update_rows_by_date(file_path, STATS_SHEET, rows_dict)
    else:
        merged = _merge(existing, fresh)
        write_stats_sheet(file_path, merged)


if __name__ == "__main__":
    import sys

    target = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "籌碼資料.xlsx"
    )
    recalculate(target)
    print(f"統計頁已更新: {target}")
